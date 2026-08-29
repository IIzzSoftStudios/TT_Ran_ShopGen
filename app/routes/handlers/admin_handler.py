"""Admin vault: registration keys and access request triage."""
from __future__ import annotations

import logging
from datetime import datetime

from io import BytesIO

from flask import (
    render_template,
    redirect,
    url_for,
    flash,
    request,
    jsonify,
    current_app,
    Response,
    send_file,
    abort,
)
from flask_login import current_user
from sqlalchemy.orm import joinedload

from app.extensions import db
from app.models import (
    AccessRequest,
    Campaign,
    CampaignCodeRedemption,
    DeletedCampaignSimSnapshot,
    ExpansionInterest,
    GMProfile,
    Player,
    PlayerCharacterSheet,
    RegistrationKey,
    User,
    UserSubmission,
)
from app.services.join_codes import mask_join_code
from app.constants.submission_categories import VALID_STATUSES
from app.services.key_generator import create_bulk_keys, generate_secure_code

audit_logger = logging.getLogger("admin_audit")


def _vault_keeper_json_forbidden():
    """If caller is not vault_keeper, return JSON Response with status 403; else None."""
    if not getattr(current_user, "is_authenticated", False):
        r = jsonify({"error": "vault_keeper_required"})
        r.status_code = 403
        return r
    if getattr(current_user, "role", None) != "vault_keeper":
        r = jsonify({"error": "vault_keeper_required"})
        r.status_code = 403
        return r
    return None


def _load_snapshot_index_by_gm() -> dict[int, list[DeletedCampaignSimSnapshot]]:
    """Group ``DeletedCampaignSimSnapshot`` rows by ``gm_profile_id``.

    A single query keeps the GM-row loop O(GMs), with O(snapshots) total work
    on the side. Snapshots have no live FK back to ``campaign``, so the
    only correctness guard we need is grouping by ``gm_profile_id``.
    """

    index: dict[int, list[DeletedCampaignSimSnapshot]] = {}
    for snap in DeletedCampaignSimSnapshot.query.order_by(
        DeletedCampaignSimSnapshot.deleted_at.asc()
    ).all():
        index.setdefault(snap.gm_profile_id, []).append(snap)
    return index


def _gm_simulation_usage_serialized_rows() -> list[dict]:
    """Read-only aggregate: one JSON-safe row per GM (simulation button click counts).

    Each row carries the GM-level aggregate (sum of click counters, sum of
    ``current_game_day - 1`` across campaigns, max ``last_tick_time``) AND a
    ``campaigns`` array with the same metrics per campaign. The vault-keeper
    UI shows the GM-level table as the default view and lets the operator
    expand a row to drill into per-campaign breakdowns.

    Deleted-campaign tombstones (``DeletedCampaignSimSnapshot``) are folded
    into both the GM-level aggregate and the per-campaign drill-down so the
    analytics view retains continuity across Campaign deletions. Tombstone
    entries are flagged ``is_deleted: True`` and carry a ``deleted_at``
    timestamp for the UI to render a "deleted" badge.
    """
    users = (
        User.query.filter(User.role.in_(["GM", "Both"]))
        .options(
            joinedload(User.gm_profile).joinedload(GMProfile.campaigns),
            joinedload(User.registration_key_used),
        )
        .order_by(User.username)
        .all()
    )
    snapshots_by_gm = _load_snapshot_index_by_gm()
    rows: list[dict] = []
    for u in users:
        gmp = u.gm_profile
        reg = u.registration_key_used
        key_phase = reg.key_phase if reg else "—"
        if gmp is None:
            rows.append(
                {
                    "username": u.username,
                    "email": u.email or "—",
                    "key_phase": key_phase,
                    "sim_clicks_day": 0,
                    "sim_clicks_week": 0,
                    "sim_clicks_month": 0,
                    "sim_clicks_year": 0,
                    "sim_clicks_pause": 0,
                    "last_tick_time": None,
                    "days_simulated": 0,
                    "campaigns_count": 0,
                    "campaigns": [],
                }
            )
            continue
        campaigns = sorted(
            gmp.campaigns or [],
            key=lambda c: ((c.name or "").lower(), c.id),
        )
        sim_clicks_day = 0
        sim_clicks_week = 0
        sim_clicks_month = 0
        sim_clicks_year = 0
        sim_clicks_pause = 0
        last_tick = None
        days_simulated = 0
        per_campaign: list[dict] = []
        for camp in campaigns:
            sim = camp.simulation_state
            c_day = int(sim.sim_clicks_day or 0) if sim is not None else 0
            c_week = int(sim.sim_clicks_week or 0) if sim is not None else 0
            c_month = int(sim.sim_clicks_month or 0) if sim is not None else 0
            c_year = int(sim.sim_clicks_year or 0) if sim is not None else 0
            c_pause = int(sim.sim_clicks_pause or 0) if sim is not None else 0
            c_last_tick = sim.last_tick_time if sim is not None else None
            c_days_simulated = max(0, int((camp.current_game_day or 1) - 1))

            sim_clicks_day += c_day
            sim_clicks_week += c_week
            sim_clicks_month += c_month
            sim_clicks_year += c_year
            sim_clicks_pause += c_pause
            if c_last_tick is not None and (
                last_tick is None or c_last_tick > last_tick
            ):
                last_tick = c_last_tick
            days_simulated += c_days_simulated

            per_campaign.append(
                {
                    "id": camp.id,
                    "name": camp.name or f"Campaign #{camp.id}",
                    "system_type": camp.system_type or "—",
                    "is_active": bool(camp.is_active),
                    "is_deleted": False,
                    "deleted_at": None,
                    "current_game_day": int(camp.current_game_day or 1),
                    "days_simulated": c_days_simulated,
                    "sim_clicks_day": c_day,
                    "sim_clicks_week": c_week,
                    "sim_clicks_month": c_month,
                    "sim_clicks_year": c_year,
                    "sim_clicks_pause": c_pause,
                    "last_tick_time": c_last_tick.isoformat() if c_last_tick else None,
                }
            )

        snapshots = snapshots_by_gm.get(gmp.id, [])
        for snap in snapshots:
            sim_clicks_day += int(snap.sim_clicks_day or 0)
            sim_clicks_week += int(snap.sim_clicks_week or 0)
            sim_clicks_month += int(snap.sim_clicks_month or 0)
            sim_clicks_year += int(snap.sim_clicks_year or 0)
            sim_clicks_pause += int(snap.sim_clicks_pause or 0)
            if snap.last_tick_time is not None and (
                last_tick is None or snap.last_tick_time > last_tick
            ):
                last_tick = snap.last_tick_time
            days_simulated += int(snap.days_simulated or 0)
            per_campaign.append(
                {
                    "id": int(snap.campaign_id),
                    "name": snap.campaign_name or f"Campaign #{snap.campaign_id}",
                    "system_type": snap.system_type or "—",
                    "is_active": False,
                    "is_deleted": True,
                    "deleted_at": snap.deleted_at.isoformat() if snap.deleted_at else None,
                    "current_game_day": int(snap.current_game_day or 1),
                    "days_simulated": int(snap.days_simulated or 0),
                    "sim_clicks_day": int(snap.sim_clicks_day or 0),
                    "sim_clicks_week": int(snap.sim_clicks_week or 0),
                    "sim_clicks_month": int(snap.sim_clicks_month or 0),
                    "sim_clicks_year": int(snap.sim_clicks_year or 0),
                    "sim_clicks_pause": int(snap.sim_clicks_pause or 0),
                    "last_tick_time": snap.last_tick_time.isoformat()
                    if snap.last_tick_time
                    else None,
                }
            )

        per_campaign.sort(
            key=lambda c: (bool(c.get("is_deleted")), (c.get("name") or "").lower(), c.get("id") or 0)
        )

        last_tick_iso = last_tick.isoformat() if last_tick else None
        rows.append(
            {
                "username": u.username,
                "email": u.email or "—",
                "key_phase": key_phase,
                "sim_clicks_day": sim_clicks_day,
                "sim_clicks_week": sim_clicks_week,
                "sim_clicks_month": sim_clicks_month,
                "sim_clicks_year": sim_clicks_year,
                "sim_clicks_pause": sim_clicks_pause,
                "last_tick_time": last_tick_iso,
                "days_simulated": days_simulated,
                "campaigns_count": len(per_campaign),
                "campaigns": per_campaign,
            }
        )
    return rows


def handle_gm_simulation_usage_api():
    denied = _vault_keeper_json_forbidden()
    if denied:
        return denied
    rows = _gm_simulation_usage_serialized_rows()
    audit_logger.info(
        "GM simulation usage API | Admin ID: %s | Rows: %s",
        current_user.id,
        len(rows),
    )
    return jsonify({"rows": rows})


def _format_iso_cell(iso: str | None) -> str:
    if not iso:
        return "—"
    try:
        if iso.endswith("Z"):
            iso = iso[:-1]
        dt = datetime.fromisoformat(iso)
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return iso


def handle_gm_simulation_usage_export():
    if not getattr(current_user, "is_authenticated", False):
        return Response("Forbidden", status=403, mimetype="text/plain")
    if getattr(current_user, "role", None) != "vault_keeper":
        return Response("Forbidden", status=403, mimetype="text/plain")

    from openpyxl import Workbook

    rows = _gm_simulation_usage_serialized_rows()
    audit_logger.info(
        "GM simulation usage export | Admin ID: %s | Rows: %s",
        current_user.id,
        len(rows),
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "GM_usage"
    headers = [
        "Username",
        "Email",
        "Key phase",
        "Run day clicks",
        "Run week clicks",
        "Run month clicks",
        "Run year clicks",
        "Pause clicks",
        "Last run (UTC)",
        "Days simulated",
        "Campaigns",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r["username"],
                r["email"],
                r["key_phase"],
                r["sim_clicks_day"],
                r["sim_clicks_week"],
                r["sim_clicks_month"],
                r["sim_clicks_year"],
                r["sim_clicks_pause"],
                _format_iso_cell(r.get("last_tick_time")),
                r.get("days_simulated", 0),
                r["campaigns_count"],
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="gm_simulation_usage.xlsx",
    )


def _parse_demo_analytics_date_bounds():
    """Optional since/until query params (YYYY-MM-DD), UTC day bounds."""
    since = None
    until = None
    since_raw = (request.args.get("since") or "").strip()
    until_raw = (request.args.get("until") or "").strip()
    if since_raw:
        try:
            since = datetime.strptime(since_raw, "%Y-%m-%d")
        except ValueError:
            since = None
    if until_raw:
        try:
            until_day = datetime.strptime(until_raw, "%Y-%m-%d")
            until = until_day.replace(hour=23, minute=59, second=59, microsecond=999999)
        except ValueError:
            until = None
    return since, until


def handle_demo_analytics_summary():
    denied = _vault_keeper_json_forbidden()
    if denied:
        return denied
    from app.services.demo_analytics import aggregate_demo_analytics

    since, until = _parse_demo_analytics_date_bounds()
    payload = aggregate_demo_analytics(since=since, until=until)
    audit_logger.info(
        "Demo analytics API | Admin ID: %s | Runs: %s",
        current_user.id,
        payload.get("total_runs"),
    )
    return jsonify(payload)


def handle_demo_analytics_export():
    if not getattr(current_user, "is_authenticated", False):
        return Response("Forbidden", status=403, mimetype="text/plain")
    if getattr(current_user, "role", None) != "vault_keeper":
        return Response("Forbidden", status=403, mimetype="text/plain")

    from openpyxl import Workbook
    from app.services.demo_analytics import aggregate_demo_analytics

    since, until = _parse_demo_analytics_date_bounds()
    payload = aggregate_demo_analytics(since=since, until=until)
    audit_logger.info(
        "Demo analytics export | Admin ID: %s | Runs: %s",
        current_user.id,
        payload.get("total_runs"),
    )
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Total demo runs", payload.get("total_runs", 0)])
    summary.append(
        ["Runs with Register click", payload.get("runs_with_register_click", 0)]
    )
    summary.append(
        ["Register conversion %", payload.get("register_conversion_pct", 0.0)]
    )

    steps_ws = wb.create_sheet("Steps")
    steps_ws.append(
        ["Step key", "Runs reached", "Reach %", "Register clicks from step"]
    )
    for row in payload.get("steps") or []:
        steps_ws.append(
            [
                row.get("step_key"),
                row.get("runs_reached", 0),
                row.get("reach_pct", 0.0),
                row.get("register_clicks", 0),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="demo_analytics.xlsx",
    )


def handle_client_analytics_summary():
    denied = _vault_keeper_json_forbidden()
    if denied:
        return denied
    from app.services.demo_analytics import aggregate_client_analytics

    since, until = _parse_demo_analytics_date_bounds()
    payload = aggregate_client_analytics(since=since, until=until)
    audit_logger.info(
        "Client analytics API | Admin ID: %s | Demo runs: %s | Submissions: %s",
        current_user.id,
        payload.get("demo_runs"),
        payload.get("submission_count"),
    )
    return jsonify(payload)


def _iso_dt(value):
    if not value:
        return None
    return value.isoformat() + "Z"


def _access_request_rows(requests):
    rows = []
    for ar in requests or []:
        rows.append(
            {
                "id": ar.id,
                "contact_name": ar.contact_name or "",
                "email": ar.email or "",
                "user_role": ar.user_role or "",
                "status": ar.status or "",
                "primary_ruleset": ar.primary_ruleset or "",
                "player_count": ar.player_count,
                "total_expected_users": ar.total_expected_users,
                "is_homebrew": bool(ar.is_homebrew),
                "discovery_source": ar.discovery_source or "",
                "notes": ar.notes or "",
                "created_at": _iso_dt(ar.created_at),
                "processed_at": _iso_dt(ar.processed_at),
            }
        )
    return rows


def _campaign_character_flat_rows(nested_rows):
    flat = []
    for user_row in nested_rows or []:
        for campaign in user_row.get("campaigns") or []:
            for player in campaign.get("players") or []:
                flat.append(
                    {
                        "user_id": user_row.get("user_id"),
                        "gm_username": user_row.get("username") or "—",
                        "gm_email": user_row.get("email") or "—",
                        "campaign_id": campaign.get("campaign_id"),
                        "campaign_name": campaign.get("campaign_name") or "—",
                        "system_type": campaign.get("system_type") or "—",
                        "is_active": bool(campaign.get("is_active")),
                        "character_name": player.get("character_name") or "—",
                        "player_username": player.get("username") or "—",
                        "player_email": player.get("email") or "—",
                        "sheet_updated_at": _iso_dt(player.get("sheet_updated_at")),
                    }
                )
    return flat


def _submission_sort_key(submission: UserSubmission):
    priority = {"pending": 0, "reviewed": 1, "closed": 2}.get(submission.status, 3)
    created = submission.created_at or datetime.min
    return (priority, -created.timestamp())


def _load_submissions_by_kind(kind: str):
    rows = UserSubmission.query.filter_by(kind=kind).all()
    rows.sort(key=_submission_sort_key)
    return rows


_CAMPAIGN_CODE_SOURCE_LABELS = {
    "registration": "Player registration",
    "registration_with_key": "Vault key + campaign code",
    "player_join": "Existing player joined",
    "campaign_selection": "Campaign selection",
}

_PROMPTED_FEEDBACK_QUESTIONS = [
    {
        "key": "campaign_limit",
        "label": "Ready to expand your realm?",
        "prompt": "You've hit the base limit for active campaigns.",
    },
    {
        "key": "monster_import",
        "label": "Import monsters",
        "prompt": "Which monster file type should import support prepare for?",
    },
    {
        "key": "srd_monsters",
        "label": "SRD monsters",
        "prompt": "Should SRD 5.1 monsters be added to the battle compendium?",
    },
    {
        "key": "market_import",
        "label": "Import market data",
        "prompt": "Which market file type should import support prepare for?",
    },
    {
        "key": "setup_tutorial",
        "label": "Setup tutorial",
        "prompt": "Would a setup tutorial be useful?",
    },
]


def _campaign_code_redemption_rows(*, limit: int = 200):
    rows = (
        CampaignCodeRedemption.query.options(
            joinedload(CampaignCodeRedemption.campaign)
            .joinedload(Campaign.gm_profile)
            .joinedload(GMProfile.user),
            joinedload(CampaignCodeRedemption.user),
        )
        .order_by(CampaignCodeRedemption.redeemed_at.desc())
        .limit(limit)
        .all()
    )
    serialized = []
    for row in rows:
        campaign = row.campaign
        gm_user = (
            campaign.gm_profile.user
            if campaign and campaign.gm_profile
            else None
        )
        serialized.append(
            {
                "code_masked": mask_join_code(
                    campaign.join_code if campaign else None
                ),
                "campaign_name": campaign.name if campaign else "—",
                "gm_username": gm_user.username if gm_user else "—",
                "player_username": row.user.username if row.user else "—",
                "source": row.source,
                "source_label": _CAMPAIGN_CODE_SOURCE_LABELS.get(
                    row.source, row.source.replace("_", " ").title()
                ),
                "redeemed_at": row.redeemed_at,
            }
        )
    return serialized


def _prompted_feedback_answer_rows(keys):
    rows_by_user = {}

    def row_for_user(user, *, key_phase="—"):
        if not user:
            return None
        row = rows_by_user.setdefault(
            user.id,
            {
                "user_id": user.id,
                "username": user.username,
                "key_phase": key_phase,
                "answers": {},
            },
        )
        if row["key_phase"] == "—" and key_phase != "—":
            row["key_phase"] = key_phase
        return row

    for key in keys:
        interest = getattr(key, "_expansion_interest", None)
        if not interest or not getattr(key, "user", None):
            continue
        latest = interest["latest"]
        row = row_for_user(key.user, key_phase=key.key_phase)
        if row is None:
            continue
        row["answers"]["campaign_limit"] = {
            "value": "No, stay base tier"
            if interest["selection"] == "no"
            else "Yes, express interest",
            "badge": "danger" if interest["selection"] == "no" else "success",
            "created_at": latest.created_at,
            "source": latest.source or "—",
        }

    prompted_keys = {q["key"] for q in _PROMPTED_FEEDBACK_QUESTIONS}
    submissions = (
        UserSubmission.query.options(
            joinedload(UserSubmission.user).joinedload(User.registration_key_used)
        )
        .filter_by(kind="suggestion")
        .order_by(UserSubmission.created_at.desc(), UserSubmission.id.desc())
        .all()
    )
    for submission in submissions:
        extra = submission.extra if isinstance(submission.extra, dict) else {}
        prompted_key = extra.get("prompted_key")
        if prompted_key not in prompted_keys or prompted_key == "campaign_limit":
            continue
        user = submission.user
        reg_key = user.registration_key_used if user else None
        row = row_for_user(
            user,
            key_phase=reg_key.key_phase if reg_key else "—",
        )
        if row is None or prompted_key in row["answers"]:
            continue
        value = extra.get("file_type") or submission.body
        row["answers"][prompted_key] = {
            "value": value,
            "badge": "secondary",
            "created_at": submission.created_at,
            "source": submission.title or "—",
        }

    return sorted(
        rows_by_user.values(),
        key=lambda row: (row["username"].lower(), row["user_id"]),
    )


def _sheet_character_name(sheet_json) -> str:
    if not isinstance(sheet_json, dict):
        return "—"
    raw = (
        sheet_json.get("name")
        or sheet_json.get("character_name")
        or sheet_json.get("display_name")
    )
    clean = str(raw or "").strip()
    return clean or "—"


def _campaign_character_rows():
    """Read-only roster grouped by user -> campaign -> players for the vault."""
    campaigns = (
        Campaign.query.options(
            joinedload(Campaign.gm_profile).joinedload(GMProfile.user),
            joinedload(Campaign.players).joinedload(Player.user),
            joinedload(Campaign.players).joinedload(Player.character_sheets),
        )
        .order_by(Campaign.name.asc(), Campaign.id.asc())
        .all()
    )
    user_rows_by_key = {}
    for campaign in campaigns:
        gm_user = (
            campaign.gm_profile.user
            if campaign.gm_profile and campaign.gm_profile.user
            else None
        )
        user_key = gm_user.id if gm_user else f"missing-gm-{campaign.gm_profile_id}"
        player_rows = []
        for player in campaign.players or []:
            if player.is_npc or player.campaign_id != campaign.id:
                continue
            sheet = next(
                (
                    row
                    for row in player.character_sheets or []
                    if row.campaign_id == campaign.id
                ),
                None,
            )
            user = player.user
            player_rows.append(
                {
                    "player_id": player.id,
                    "username": user.username if user else "—",
                    "email": (user.email if user and user.email else "—"),
                    "character_name": _sheet_character_name(
                        sheet.sheet_json if sheet else None
                    ),
                    "sheet_updated_at": sheet.updated_at if sheet else None,
                }
            )
        player_rows.sort(
            key=lambda row: (
                (row["character_name"] if row["character_name"] != "—" else "").lower(),
                row["username"].lower(),
                row["player_id"] or 0,
            )
        )
        user_row = user_rows_by_key.setdefault(
            user_key,
            {
                "user_id": gm_user.id if gm_user else None,
                "username": gm_user.username if gm_user else "—",
                "email": gm_user.email if gm_user and gm_user.email else "—",
                "campaign_count": 0,
                "player_count": 0,
                "campaigns": [],
            },
        )
        user_row["campaigns"].append(
            {
                "campaign_id": campaign.id,
                "campaign_name": campaign.name or f"Campaign #{campaign.id}",
                "system_type": campaign.system_type or "—",
                "is_active": bool(campaign.is_active),
                "player_count": len(player_rows),
                "players": player_rows,
            }
        )
        user_row["campaign_count"] += 1
        user_row["player_count"] += len(player_rows)

    rows = list(user_rows_by_key.values())
    for user_row in rows:
        user_row["campaigns"].sort(
            key=lambda row: (
                0 if row["is_active"] else 1,
                row["campaign_name"].lower(),
                row["campaign_id"] or 0,
            )
        )
    rows.sort(
        key=lambda row: (
            row["username"].lower(),
            row["user_id"] or 0,
        )
    )
    return rows


def handle_admin_keys():
    keys = (
        RegistrationKey.query.filter_by(is_admin_test_key=False)
        .order_by(RegistrationKey.created_at.desc())
        .all()
    )
    _attach_expansion_interest_metadata(keys)
    admin_keys = (
        RegistrationKey.query.filter_by(is_admin_test_key=True)
        .order_by(RegistrationKey.created_at.desc())
        .all()
    )
    stats = {
        "total": len(keys),
        "used": sum(1 for k in keys if k.is_used),
        "available": sum(1 for k in keys if not k.is_used),
    }
    admin_stats = {
        "total": len(admin_keys),
        "used": sum(1 for k in admin_keys if k.is_used),
        "available": sum(1 for k in admin_keys if not k.is_used),
    }
    access_requests = AccessRequest.query.all()
    role_order = {"GM": 0, "Both": 1, "Player": 2}

    def sort_key(r):
        status_priority = 0 if r.status == "pending" else 1
        return (
            status_priority,
            r.primary_ruleset or "",
            role_order.get(r.user_role, 99),
            r.queue_sort_ts or r.created_at,
        )

    access_requests.sort(key=sort_key)
    prompted_feedback_rows = _prompted_feedback_answer_rows(keys)
    prompted_feedback_questions = _PROMPTED_FEEDBACK_QUESTIONS
    pc = current_app.extensions["phase_config"]
    vault_phase_slugs = pc.list_phases(include_internal=False)
    all_phase_slugs = pc.list_phases(include_internal=True)
    show_gm_usage_tab = getattr(current_user, "role", None) == "vault_keeper"
    gm_simulation_rows = (
        _gm_simulation_usage_serialized_rows() if show_gm_usage_tab else []
    )
    demo_analytics = None
    demo_funnel_steps = []
    demo_lead_step = None
    demo_leads_at_step = []
    if show_gm_usage_tab:
        from app.services.demo_analytics import aggregate_demo_analytics
        from app.services.demo_leads import funnel_step_options, leads_stopped_at

        demo_analytics = aggregate_demo_analytics()
        demo_funnel_steps = funnel_step_options()
        demo_lead_step = (request.args.get("demo_step") or "").strip() or None
        demo_leads_at_step = leads_stopped_at(demo_lead_step) if demo_lead_step else []
        from app.services.demo_analytics import aggregate_client_analytics

        client_analytics = aggregate_client_analytics()
    else:
        client_analytics = None
    access_request_rows = _access_request_rows(access_requests)
    campaign_code_redemptions = _campaign_code_redemption_rows()
    stats["total"] += len(campaign_code_redemptions)
    stats["used"] += len(campaign_code_redemptions)
    campaign_character_rows = _campaign_character_rows()
    campaign_character_flat_rows = _campaign_character_flat_rows(campaign_character_rows)
    audit_logger.info("Keys view | Admin ID: %s", current_user.id)
    return render_template(
        "admin/keys.html",
        keys=keys,
        admin_keys=admin_keys,
        stats=stats,
        admin_stats=admin_stats,
        access_requests=access_requests,
        prompted_feedback_rows=prompted_feedback_rows,
        prompted_feedback_questions=prompted_feedback_questions,
        vault_phase_slugs=vault_phase_slugs,
        all_phase_slugs=all_phase_slugs,
        show_gm_usage_tab=show_gm_usage_tab,
        gm_simulation_rows=gm_simulation_rows,
        demo_analytics=demo_analytics,
        demo_funnel_steps=demo_funnel_steps,
        demo_lead_step=demo_lead_step,
        demo_leads_at_step=demo_leads_at_step,
        client_analytics=client_analytics,
        access_request_rows=access_request_rows,
        campaign_code_redemptions=campaign_code_redemptions,
        campaign_character_rows=campaign_character_rows,
        campaign_character_flat_rows=campaign_character_flat_rows,
        bug_reports=_load_submissions_by_kind("bug_report"),
        feedback_items=_load_submissions_by_kind("feedback"),
        suggestions=_load_submissions_by_kind("suggestion"),
    )


def _attach_expansion_interest_metadata(keys):
    user_ids = sorted({k.user_id for k in keys if getattr(k, "user_id", None)})
    if not user_ids:
        for key in keys:
            key._expansion_interest = None
        return

    latest_rows = (
        ExpansionInterest.query.filter(ExpansionInterest.user_id.in_(user_ids))
        .order_by(
            ExpansionInterest.user_id.asc(),
            ExpansionInterest.created_at.desc(),
            ExpansionInterest.id.desc(),
        )
        .all()
    )
    latest_by_user = {}
    for row in latest_rows:
        latest_by_user.setdefault(row.user_id, row)

    for key in keys:
        user_id = getattr(key, "user_id", None)
        latest = latest_by_user.get(user_id)
        key._expansion_interest = (
            {
                "selection": "no" if latest.intent == "not_interested" else "yes",
                "latest": latest,
            }
            if latest is not None
            else None
        )


def handle_submission_action(submission_id: int, action: str):
    if getattr(current_user, "role", None) != "vault_keeper":
        abort(404)
    submission = UserSubmission.query.get_or_404(submission_id)
    if action == "review":
        if submission.status != "pending":
            return (
                jsonify(
                    {
                        "error": (
                            "Only pending submissions can be marked reviewed."
                        )
                    }
                ),
                409,
            )
        submission.status = "reviewed"
    elif action == "close":
        if submission.status == "closed":
            return jsonify({"error": "Submission is already archived."}), 409
        submission.status = "closed"
    else:
        return jsonify({"error": "Unknown action."}), 400
    if submission.status not in VALID_STATUSES:
        return jsonify({"error": "Invalid status."}), 400
    db.session.commit()
    audit_logger.info(
        "Submission %s | id=%s | admin=%s",
        action,
        submission_id,
        current_user.id,
    )
    return jsonify({"success": True, "new_status": submission.status})


def _validate_phase_slug(selected: str | None) -> str | None:
    """Return normalized slug or None if invalid."""
    if not selected or not str(selected).strip():
        return None
    slug = str(selected).strip()
    pc = current_app.extensions["phase_config"]
    if slug not in pc.list_phases(include_internal=True):
        return None
    return slug


def handle_generate_bulk():
    selected_phase = _validate_phase_slug(request.form.get("key_phase"))
    if selected_phase is None:
        flash("Invalid phase selection.", "error")
        return redirect(url_for("admin.keys_overview"))
    try:
        count = int(request.form.get("count", 5))
    except (TypeError, ValueError):
        count = 5
    count = max(1, min(50, count))
    try:
        create_bulk_keys(count, email=None, is_admin_test_key=False, phase_slug=selected_phase)
    except ValueError:
        flash("Invalid phase selection.", "error")
        return redirect(url_for("admin.keys_overview"))
    db.session.commit()
    audit_logger.info(
        "Keys generated | Admin ID: %s | Count: %s | Phase: %s",
        current_user.id,
        count,
        selected_phase,
    )
    flash(f"Generated {count} new keys ({selected_phase}).", "success")
    return redirect(url_for("admin.keys_overview"))


def handle_generate_admin_test_keys():
    try:
        count = int(request.form.get("count", 3))
    except (TypeError, ValueError):
        count = 3
    count = max(1, min(20, count))
    create_bulk_keys(count, email=None, is_admin_test_key=True)
    db.session.commit()
    audit_logger.info(
        "Admin test keys generated | Admin ID: %s | Count: %s", current_user.id, count
    )
    flash(f"Generated {count} admin test key(s).", "success")
    return redirect(url_for("admin.keys_overview"))


def _generate_unique_vault_key(phase_slug: str) -> str:
    pc = current_app.extensions["phase_config"]
    row = pc.get_phase(phase_slug)
    prefix = row["prefix"]
    while True:
        code = generate_secure_code(prefix=prefix, segments=2, segment_len=4)
        exists = RegistrationKey.query.filter_by(key_code=code).first() or AccessRequest.query.filter_by(
            vault_key=code
        ).first()
        if not exists:
            return code


def handle_approve_access_request(request_id):
    ar = AccessRequest.query.get_or_404(request_id)
    selected_phase = _validate_phase_slug(request.form.get("key_phase"))
    if selected_phase is None:
        flash("Invalid phase selection.", "error")
        return redirect(url_for("admin.keys_overview"))

    ar.status = "approved"
    vault_key = _generate_unique_vault_key(selected_phase)
    ar.vault_key = vault_key
    ar.vault_key_used = False
    ar.processed_at = datetime.utcnow()
    reg_key = RegistrationKey(
        key_code=vault_key,
        email=ar.email,
        is_used=False,
        key_phase=selected_phase,
        is_admin_test_key=False,
    )
    db.session.add(reg_key)
    db.session.commit()

    try:
        from flask_mailman import EmailMessage

        sender = current_app.config.get("MAIL_DEFAULT_SENDER", "noreply@example.com")
        reg_link = url_for("main.register_alias", vault_key=vault_key, _external=True)
        subject = "Econo-Forge Access Approved"
        discord_link = "https://discord.gg/dkNVnBXMMB"
        body = (
            f"Your Econo-Forge registration key was issued by an admin review.\n\n"
            f"Registration key:\n{vault_key}\n\n"
            f"Complete registration:\n{reg_link}\n\n"
            f"Optional community updates: {discord_link}\n"
        )
        msg = EmailMessage(subject, body, sender, [ar.email])
        msg.send()
    except Exception as e:
        audit_logger.warning("Approval email failed | AccessRequest ID: %s | Error: %s", ar.id, e)

    flash("Access request approved. Registration key issued.", "success")
    return redirect(url_for("admin.keys_overview"))


def handle_hold_access_request(request_id):
    ar = AccessRequest.query.get_or_404(request_id)
    ar.status = "hold"
    ar.queue_sort_ts = datetime.utcnow()
    db.session.commit()
    flash("Access request held (moved to bottom).", "info")
    return redirect(url_for("admin.keys_overview"))


def handle_reject_access_request(request_id):
    ar = AccessRequest.query.get_or_404(request_id)
    ar.status = "rejected"
    ar.processed_at = datetime.utcnow()
    db.session.commit()
    flash("Access request rejected.", "warning")
    return redirect(url_for("admin.keys_overview"))


def handle_reveal_key(key_id):
    key_row = RegistrationKey.query.get_or_404(key_id)
    audit_logger.info("Key Reveal | Admin ID: %s | Key ID: %s", current_user.id, key_id)
    if key_row.is_used:
        return jsonify({"error": "Key already used"}), 400
    return jsonify({"key_code": key_row.key_code})
